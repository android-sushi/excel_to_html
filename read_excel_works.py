import openpyxl as xl
import pathlib
import re
from datetime import datetime


def read_excel_works():
    dir_path = pathlib.Path('Excel作業内容/2021/04/01')

    excel_data = {}
    excel_data_value = {}
    excel_row = []
    work_content = {}
    person_name = []

    for path in dir_path.iterdir():
        file_name = re.split('[_.]', path.name)
        company_name = file_name[1]
        date = datetime.strptime(file_name[0], '%Y%m%d')
        excel_data_value['日付'] = date

        wb = xl.load_workbook(path)
        ws = wb['作業内容']
        for row in ws.iter_rows(min_row=2):
            place = row[0].value
            content = row[1].value
            classifying = row[2].value
            excel_row.append([place, content, classifying])
        excel_data_value['作業内容'] = excel_row
        excel_row = []

        ws = wb['出勤名簿']
        for row in ws.iter_rows(min_row=2):
            if row[0].value == 1:
                person_name.append([row[i].value for i in range(1,11)])
        excel_data_value['出勤名簿'] = person_name
        person_name = []

        excel_data[company_name] = excel_data_value
        excel_data_value = {}    

    return excel_data


if __name__ == '__main__':
    excel_data = read_excel_works()
    for keys, values in excel_data.items():
        print(keys)
        for key, value in values.items():
            print(key, value)
        print()